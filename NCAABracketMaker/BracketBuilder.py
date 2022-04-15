from NCAABracketMaker.utilities import bracketpath, simbracketpath, teampath, modulepath
import yaml
from NCAABracketMaker.AnalyzeGame import whoWins
from math import pow
import pandas as pd
import os
from openpyxl import load_workbook


def roundResults(teams, rnd, teamdatadf, pointcof, wincof, rankcof, ratiocof):
    """
    Determins the winning teams from each region for the given round of games
    :param teams:
    :param rnd: int, current round of games
    :param teamdatadf: dataframe, all team data
    :param pointcof:  float, points weight coefficient
    :param wincof: float, number of wins weight coefficient
    :param rankcof: float, top 25 team rank weight coefficient
    :param ratiocof: float, wins to losses weight coefficient
    :return: dataframe, list of winning teams
    """
    # Tests each game for defined round
    winners = {}

    # Loops through each game by region
    for x in range(1, 5):
        for y in range(1, int(pow(2, 4 - rnd) + 1)):
            seedid = f"d{rnd}r{x}seed"
            winners[f"d{rnd + 1}r{x}seed{y}"] = whoWins(
                teams.get("".join([seedid, str(y)])),
                teams.get("".join([seedid, str(int(pow(2, 5 - rnd) + 1 - y))])),
                teamdatadf,
                pointcof,
                wincof,
                rankcof,
                ratiocof,
            )

    return winners


def bracketSim(bracketfile, teamdatafile, pointcof, wincof, rankcof, ratiocof):
    """
    Simulates the results of a bracket
    :param bracketfile: string, file for bracket and teams the program is simulating
    :param teamdatafile: string, filepath for csv of all team data
    :param pointcof:  float, points weight coefficient
    :param wincof: float, number of wins weight coefficient
    :param rankcof: float, top 25 team rank weight coefficient
    :param ratiocof: float, wins to losses weight coefficient
    :return: None, saves a csv with the results in SimBrackets directory
    """
    if "2020" in bracketfile:
        # 2020 tournament didn't happen so creates an empty file
        with open(
            f'{simbracketpath}{teamdatafile.replace(".csv", "")}Sim.csv', "w"
        ) as f:
            f.write("Coronavirus")
        print("Coronavirus")
        return None

    teamdatadf = pd.read_csv(teampath + teamdatafile, index_col="Team Name")
    # Simulates all games in the supplied brackets based on teamdatafile info
    with open(bracketpath + bracketfile) as f:
        bracketData = yaml.load(f, Loader=yaml.FullLoader)

    data = {k: v for k, v in bracketData.items() if k[:2] == "d1"}

    # First 4 games, different years have different first 4 games
    first4seeds = {
        "d1r1seed11",
        "d1r2seed11",
        "d1r3seed11",
        "d1r4seed11",
        "d1r4seed12",
        "d1r1seed16",
        "d1r2seed16",
        "d1r3seed16",
        "d1r4seed16",
        "d1r4seed11",
        "d1r2seed11",
        "d1r4seed12",
    }
    for i in first4seeds:
        try:
            data[i] = whoWins(
                data.get(f"{i}a"),
                data.get(f"{i}b"),
                teamdatadf,
                pointcof,
                wincof,
                rankcof,
                ratiocof,
            )
        except (LookupError, ValueError):
            pass

    # Simulates and stores each round until winner is determined
    champion = {}
    round1winners = roundResults(
        data, 1, teamdatadf, pointcof, wincof, rankcof, ratiocof
    )
    round2winners = roundResults(
        round1winners, 2, teamdatadf, pointcof, wincof, rankcof, ratiocof
    )
    round3winners = roundResults(
        round2winners, 3, teamdatadf, pointcof, wincof, rankcof, ratiocof
    )
    round4winners = roundResults(
        round3winners, 4, teamdatadf, pointcof, wincof, rankcof, ratiocof
    )
    round5winners = {
        "d6r1seed1": whoWins(
            round4winners.get("d5r1seed1"),
            round4winners.get("d5r4seed1"),
            teamdatadf,
            pointcof,
            wincof,
            rankcof,
            ratiocof,
        ),
        "d6r1seed2": whoWins(
            round4winners.get("d5r2seed1"),
            round4winners.get("d5r3seed1"),
            teamdatadf,
            pointcof,
            wincof,
            rankcof,
            ratiocof,
        ),
    }
    champion["d7r1seed1"] = whoWins(
        round5winners.get("d6r1seed1"),
        round5winners.get("d6r1seed2"),
        teamdatadf,
        pointcof,
        wincof,
        rankcof,
        ratiocof,
    )

    try:
        data.pop("d1r1seed16a")
        data.pop("d1r1seed16b")
        data.pop("d1r1seed12a")
        data.pop("d1r1seed12b")
        data.pop("d1r4seed16a")
        data.pop("d1r4seed16b")
        data.pop("d1r4seed12a")
        data.pop("d1r4seed12b")
    except KeyError:
        pass

    totalsimData = {
        **data,
        **round1winners,
        **round2winners,
        **round3winners,
        **round4winners,
        **round5winners,
        **champion,
    }

    # ensure directory exists
    os.makedirs(f"{modulepath}SimBrackets/", exist_ok=True)
    datafile = teamdatafile.replace(".csv", "")

    with open(
        f"{simbracketpath}{datafile}-{pointcof}-{wincof}-{rankcof}-{ratiocof}-Sim.csv",
        "w",
    ) as f:
        for key in totalsimData.keys():
            f.write(f"{key}, {totalsimData[key]}\n")

    return None


def populateBracket(simbracket):
    """
    Adds the simulated bracket results to xlsx file to make it easier to read and compare results
    :param simbracket: string, filepath for csv of simulate bracket results
    :return: None, saves xlsx file with simbracket results
    """
    # Creates dataframe with simbracket
    df = pd.read_csv(f"{simbracketpath}{simbracket}")
    filename = simbracket.rstrip(".csv")
    #  Adds sheet with simbracket data
    with pd.ExcelWriter(
        f"{simbracketpath}Sim_Bracket.xlsx", mode="a", if_sheet_exists="replace"
    ) as writer:
        df.to_excel(writer, sheet_name=filename)

    # Opens Sim_Bracket.xlsx and adds sheet name to sheet list
    workbook = load_workbook(filename=f"{simbracketpath}Sim_Bracket.xlsx")
    sheet = workbook["Bracket"]

    # Checks for other sheets listed
    for x in range(5, 21):
        if sheet[f"AE{x}"].value is None:
            sheet[f"AE{x}"] = filename
            break
    # Saves workbook
    workbook.save(filename=f"{simbracketpath}Sim_Bracket.xlsx")

    return None


def clearSimResults(reset):
    """
    Deletes all csv files in SimBrackets and clears results from Sim_Bracket.xlsx
    :param reset: Bool, defaults False, True deletes csv files and clears Sim_Bracket.xlsx
    :return: None
    """
    if reset:
        # Opens Sim_Bracket.xlsx and deletes sheets
        workbook = load_workbook(filename=f"{simbracketpath}Sim_Bracket.xlsx")
        sheets = workbook.sheetnames
        for i in sheets:
            if i == "Bracket":
                pass
            else:
                workbook.remove(workbook[i])

        # Remove sheets from drop down list
        sheet = workbook["Bracket"]
        for x in range(5, 21):
            sheet[f"AE{x}"].value = None

        # Saves workbook
        workbook.save(filename=f"{simbracketpath}Sim_Bracket.xlsx")

        # Deletes all csv files in SimBrackets directory
        for f in os.listdir(simbracketpath):
            if not f.endswith(".csv"):
                continue
            os.remove(os.path.join(simbracketpath, f))

    return None
